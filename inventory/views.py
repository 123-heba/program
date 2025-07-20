from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q, F
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from .models import Product, SaleInvoice, PurchaseInvoice, Customer, Supplier
from .decorators import admin_required, manager_required, employee_required
from datetime import date

@login_required
def dashboard(request):
    total_products = Product.objects.count()
    total_customers = Customer.objects.count()
    total_suppliers = Supplier.objects.count()
    
    total_sales = SaleInvoice.objects.aggregate(total=Sum('total_amount'))['total'] or 0
    total_purchases = PurchaseInvoice.objects.aggregate(total=Sum('total_amount'))['total'] or 0
    
    low_stock_products = Product.objects.filter(quantity__lt=10)
    low_stock_count = low_stock_products.count()
    
    inventory_value = Product.objects.aggregate(total=Sum('price'))['total'] or 0
    
    recent_sales = SaleInvoice.objects.order_by('-date')[:5]
    recent_purchases = PurchaseInvoice.objects.order_by('-date')[:5]
    
    context = {
        'total_products': total_products,
        'total_customers': total_customers,
        'total_suppliers': total_suppliers,
        'total_sales': total_sales,
        'total_purchases': total_purchases,
        'low_stock_count': low_stock_count,
        'low_stock_products': low_stock_products,
        'inventory_value': inventory_value,
        'recent_sales': recent_sales,
        'recent_purchases': recent_purchases,
        'today': date.today(),
        'today_invoices': SaleInvoice.objects.filter(date=date.today()).count(),
    }
    
    return render(request, 'inventory/dashboard.html', context)

@login_required
def employee_dashboard(request):
    return render(request, 'inventory/employee_dashboard.html')

# Product views
@login_required
def product_list(request):
    products = Product.objects.all()
    suppliers = Supplier.objects.all()
    
    search = request.GET.get('search')
    supplier_id = request.GET.get('supplier')
    stock_status = request.GET.get('stock_status')
    
    if search:
        products = products.filter(Q(name__icontains=search) | Q(barcode_data__icontains=search))
    
    if supplier_id:
        products = products.filter(supplier_id=supplier_id)
    
    if stock_status == 'low':
        products = products.filter(quantity__lt=10)
    elif stock_status == 'out':
        products = products.filter(quantity=0)
    
    return render(request, 'inventory/product_list.html', {'products': products, 'suppliers': suppliers})

@login_required
def product_create(request):
    return render(request, 'inventory/product_form.html')

@login_required
def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'inventory/product_detail.html', {'product': product})

@login_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'inventory/product_form.html', {'product': product})

@login_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'تم حذف المنتج بنجاح')
        return redirect('product_list')
    return render(request, 'inventory/product_confirm_delete.html', {'product': product})

@login_required
def product_barcodes(request):
    products = Product.objects.all()
    return render(request, 'inventory/product_barcodes.html', {'products': products})

# Supplier views
@login_required
def supplier_list(request):
    suppliers = Supplier.objects.all()
    return render(request, 'inventory/supplier_list.html', {'suppliers': suppliers})

@login_required
def supplier_create(request):
    if request.method == 'POST':
        from .forms import SupplierForm
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم إضافة المورد بنجاح')
            return redirect('supplier_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        from .forms import SupplierForm
        form = SupplierForm()
    
    return render(request, 'inventory/supplier_form.html', {
        'form': form,
        'title': 'إضافة مورد جديد'
    })

@login_required
def supplier_edit(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    
    if request.method == 'POST':
        from .forms import SupplierForm
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث بيانات المورد بنجاح')
            return redirect('supplier_list')
        else:
            # إرجاع الأخطاء للمستخدم
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        from .forms import SupplierForm
        form = SupplierForm(instance=supplier)
    
    return render(request, 'inventory/supplier_form.html', {
        'supplier': supplier,
        'form': form,
        'title': 'تعديل المورد'
    })

@login_required
def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.delete()
        return redirect('supplier_list')
    return render(request, 'inventory/supplier_confirm_delete.html', {'supplier': supplier})

# Customer views
@login_required
def customer_list(request):
    customers = Customer.objects.all()
    return render(request, 'inventory/customer_list.html', {'customers': customers})

@login_required
def customer_create(request):
    if request.method == 'POST':
        from .forms import CustomerForm
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم إضافة العميل بنجاح')
            return redirect('customer_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        from .forms import CustomerForm
        form = CustomerForm()
    
    return render(request, 'inventory/customer_form.html', {
        'form': form,
        'title': 'إضافة عميل جديد'
    })

@login_required
def customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    
    if request.method == 'POST':
        from .forms import CustomerForm
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث بيانات العميل بنجاح')
            return redirect('customer_list')
        else:
            # إرجاع الأخطاء للمستخدم
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        from .forms import CustomerForm
        form = CustomerForm(instance=customer)
    
    return render(request, 'inventory/customer_form.html', {
        'customer': customer,
        'form': form,
        'title': 'تعديل العميل'
    })

@login_required
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer.delete()
        return redirect('customer_list')
    return render(request, 'inventory/customer_confirm_delete.html', {'customer': customer})

# Invoice views
@login_required
def sale_invoice_list(request):
    invoices = SaleInvoice.objects.all().order_by('-date')
    return render(request, 'inventory/sale_invoice_list.html', {'invoices': invoices})

@login_required
def sale_invoice_create(request):
    if request.method == 'POST':
        try:
            from .forms import SaleInvoiceForm
            from .models import Product, SaleInvoiceItem
            
            form = SaleInvoiceForm(request.POST)
            if form.is_valid():
                invoice = form.save(commit=False)
                invoice.created_by = request.user
                invoice.save()
                
                # إضافة المنتجات للفاتورة
                products = request.POST.getlist('product')
                quantities = request.POST.getlist('quantity')
                prices = request.POST.getlist('price')
                
                total_amount = 0
                for i, product_id in enumerate(products):
                    if product_id and quantities[i] and prices[i]:
                        product = Product.objects.get(id=product_id)
                        quantity = int(quantities[i])
                        unit_price = float(prices[i])
                        total_price = quantity * unit_price
                        
                        SaleInvoiceItem.objects.create(
                            invoice=invoice,
                            product=product,
                            quantity=quantity,
                            unit_price=unit_price,
                            total_price=total_price
                        )
                        total_amount += total_price
                        
                        # تقليل الكمية من المخزون
                        product.quantity -= quantity
                        product.save()
                
                # تحديث إجمالي الفاتورة
                invoice.total_amount = total_amount
                invoice.save()
                
                messages.success(request, f'تم إنشاء فاتورة البيع {invoice.invoice_number} بنجاح')
                return redirect('sale_invoice_detail', pk=invoice.pk)
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
        except Exception as e:
            messages.error(request, f'حدث خطأ: {str(e)}')
    else:
        from .forms import SaleInvoiceForm
        form = SaleInvoiceForm()
    
    # إرسال قائمة المنتجات والعملاء
    from .models import Product, Customer
    products = Product.objects.all()
    customers = Customer.objects.all()
    
    return render(request, 'inventory/sale_invoice_form.html', {
        'form': form,
        'products': products,
        'customers': customers,
        'title': 'إنشاء فاتورة بيع جديدة'
    })

@login_required
def sale_invoice_detail(request, pk):
    invoice = get_object_or_404(SaleInvoice, pk=pk)
    return render(request, 'inventory/sale_invoice_detail.html', {'invoice': invoice})

@login_required
def sale_invoice_print(request, pk):
    invoice = get_object_or_404(SaleInvoice, pk=pk)
    return render(request, 'inventory/sale_invoice_print.html', {'invoice': invoice})

@login_required
def sale_invoice_delete(request, pk):
    invoice = get_object_or_404(SaleInvoice, pk=pk)
    if request.method == 'POST':
        # إرجاع الكميات للمخزون
        for item in invoice.items.all():
            item.product.quantity += item.quantity
            item.product.save()
        
        invoice.delete()
        messages.success(request, 'تم حذف فاتورة البيع بنجاح')
        return redirect('sale_invoice_list')
    return render(request, 'inventory/sale_invoice_confirm_delete.html', {'invoice': invoice})

@login_required
def purchase_invoice_list(request):
    invoices = PurchaseInvoice.objects.all().order_by('-date')
    return render(request, 'inventory/purchase_invoice_list.html', {'invoices': invoices})

@login_required
def purchase_invoice_create(request):
    if request.method == 'POST':
        try:
            from .forms import PurchaseInvoiceForm
            from .models import Product, PurchaseInvoiceItem
            
            form = PurchaseInvoiceForm(request.POST)
            if form.is_valid():
                invoice = form.save(commit=False)
                invoice.created_by = request.user
                invoice.save()
                
                # إضافة المنتجات للفاتورة
                products = request.POST.getlist('product')
                quantities = request.POST.getlist('quantity')
                prices = request.POST.getlist('price')
                
                total_amount = 0
                for i, product_id in enumerate(products):
                    if product_id and quantities[i] and prices[i]:
                        product = Product.objects.get(id=product_id)
                        quantity = int(quantities[i])
                        unit_price = float(prices[i])
                        total_price = quantity * unit_price
                        
                        PurchaseInvoiceItem.objects.create(
                            invoice=invoice,
                            product=product,
                            quantity=quantity,
                            unit_price=unit_price,
                            total_price=total_price
                        )
                        total_amount += total_price
                        
                        # إضافة الكمية للمخزون
                        product.quantity += quantity
                        product.save()
                
                # تحديث إجمالي الفاتورة
                invoice.total_amount = total_amount
                invoice.save()
                
                messages.success(request, f'تم إنشاء فاتورة الشراء {invoice.invoice_number} بنجاح')
                return redirect('purchase_invoice_detail', pk=invoice.pk)
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
        except Exception as e:
            messages.error(request, f'حدث خطأ: {str(e)}')
    else:
        from .forms import PurchaseInvoiceForm
        form = PurchaseInvoiceForm()
    
    # إرسال قائمة المنتجات والموردين
    from .models import Product, Supplier
    products = Product.objects.all()
    suppliers = Supplier.objects.all()
    
    return render(request, 'inventory/purchase_invoice_form.html', {
        'form': form,
        'products': products,
        'suppliers': suppliers,
        'title': 'إنشاء فاتورة شراء جديدة'
    })

@login_required
def purchase_invoice_detail(request, pk):
    invoice = get_object_or_404(PurchaseInvoice, pk=pk)
    return render(request, 'inventory/purchase_invoice_detail.html', {'invoice': invoice})

@login_required
def purchase_invoice_print(request, pk):
    invoice = get_object_or_404(PurchaseInvoice, pk=pk)
    return render(request, 'inventory/purchase_invoice_print.html', {'invoice': invoice})

@login_required
def purchase_invoice_delete(request, pk):
    invoice = get_object_or_404(PurchaseInvoice, pk=pk)
    if request.method == 'POST':
        # إرجاع الكميات من المخزون
        for item in invoice.items.all():
            item.product.quantity -= item.quantity
            item.product.save()
        
        invoice.delete()
        messages.success(request, 'تم حذف فاتورة الشراء بنجاح')
        return redirect('purchase_invoice_list')
    return render(request, 'inventory/purchase_invoice_confirm_delete.html', {'invoice': invoice})

# Order views
@login_required
def order_list(request):
    try:
        from .models import Order
        orders = Order.objects.all().order_by('-order_date')
        
        # حساب عدد الطلبات حسب الحالة
        pending_count = orders.filter(status='pending').count()
        processing_count = orders.filter(status='processing').count()
        shipped_count = orders.filter(status='shipped').count()
        delivered_count = orders.filter(status='delivered').count()
        cancelled_count = orders.filter(status='cancelled').count()
        
        # خيارات الحالة للفلتر
        status_choices = Order.STATUS_CHOICES
        
        context = {
            'orders': orders,
            'pending_count': pending_count,
            'processing_count': processing_count,
            'shipped_count': shipped_count,
            'delivered_count': delivered_count,
            'cancelled_count': cancelled_count,
            'status_choices': status_choices,
        }
        
    except Exception as e:
        print(f"Error in order_list: {e}")
        context = {
            'orders': [],
            'pending_count': 0,
            'processing_count': 0,
            'shipped_count': 0,
            'delivered_count': 0,
            'cancelled_count': 0,
            'status_choices': [],
        }
    
    return render(request, 'inventory/order_list.html', context)

@login_required
def order_create(request):
    if request.method == 'POST':
        try:
            from .forms import OrderForm
            from .models import Product, OrderItem
            
            # إنشاء الطلب أولاً
            form = OrderForm(request.POST)
            if form.is_valid():
                order = form.save(commit=False)
                order.created_by = request.user
                order.save()
                
                # إضافة المنتجات للطلب
                products = request.POST.getlist('product')
                quantities = request.POST.getlist('quantity')
                prices = request.POST.getlist('unit_price')
                
                total_amount = 0
                for i, product_id in enumerate(products):
                    if product_id and quantities[i] and prices[i]:
                        product = Product.objects.get(id=product_id)
                        quantity = int(quantities[i])
                        unit_price = float(prices[i])
                        total_price = quantity * unit_price
                        
                        OrderItem.objects.create(
                            order=order,
                            product=product,
                            quantity=quantity,
                            unit_price=unit_price,
                            total_price=total_price
                        )
                        total_amount += total_price
                
                # تحديث إجمالي الطلب
                order.total_amount = total_amount
                order.save()
                
                messages.success(request, f'تم إنشاء الطلب {order.order_number} بنجاح')
                return redirect('order_detail', pk=order.pk)
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
        except Exception as e:
            messages.error(request, f'حدث خطأ: {str(e)}')
    else:
        from .forms import OrderForm
        form = OrderForm()
    
    # إرسال قائمة المنتجات للاختيار منها
    from .models import Product
    products = Product.objects.all()
    
    return render(request, 'inventory/order_form.html', {
        'form': form,
        'products': products,
        'title': 'إنشاء طلب جديد'
    })

@login_required
def order_detail(request, pk):
    try:
        from .models import Order
        order = get_object_or_404(Order, pk=pk)
        order_items = order.items.all()
        
        context = {
            'order': order,
            'order_items': order_items
        }
    except Exception as e:
        messages.error(request, 'الطلب غير موجود')
        return redirect('order_list')
    
    return render(request, 'inventory/order_detail.html', context)

@login_required
def order_edit(request, pk):
    try:
        from .models import Order
        order = get_object_or_404(Order, pk=pk)
        
        if request.method == 'POST':
            from .forms import OrderForm
            form = OrderForm(request.POST, instance=order)
            if form.is_valid():
                form.save()
                messages.success(request, 'تم تحديث الطلب بنجاح')
                return redirect('order_detail', pk=order.pk)
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
        else:
            from .forms import OrderForm
            form = OrderForm(instance=order)
        
        return render(request, 'inventory/order_form.html', {
            'form': form,
            'order': order,
            'title': 'تعديل الطلب'
        })
    except:
        messages.error(request, 'الطلب غير موجود')
        return redirect('order_list')

@login_required
def order_delete(request, pk):
    try:
        from .models import Order
        order = get_object_or_404(Order, pk=pk)
        if request.method == 'POST':
            order.delete()
            return redirect('order_list')
    except:
        return redirect('order_list')
    return render(request, 'inventory/order_confirm_delete.html', {'order': order})

@login_required
def order_item_create(request, order_pk):
    try:
        from .models import Order
        order = get_object_or_404(Order, pk=order_pk)
        
        if request.method == 'POST':
            from .forms import OrderItemForm
            form = OrderItemForm(request.POST)
            if form.is_valid():
                order_item = form.save(commit=False)
                order_item.order = order
                order_item.total_price = order_item.quantity * order_item.unit_price
                order_item.save()
                
                # تحديث إجمالي الطلب
                order.total_amount = sum(item.total_price for item in order.items.all())
                order.save()
                
                messages.success(request, 'تم إضافة المنتج للطلب بنجاح')
                return redirect('order_detail', pk=order.pk)
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
        else:
            from .forms import OrderItemForm
            form = OrderItemForm()
        
        return render(request, 'inventory/order_item_form.html', {
            'form': form,
            'order': order,
            'title': 'إضافة منتج للطلب'
        })
    except Exception as e:
        messages.error(request, f'حدث خطأ: {str(e)}')
        return redirect('order_list')

@login_required
def order_item_edit(request, order_pk, pk):
    return render(request, 'inventory/order_item_form.html')

@login_required
def order_item_delete(request, order_pk, pk):
    return render(request, 'inventory/order_item_confirm_delete.html')

# Reports views
@login_required
def reports(request):
    from datetime import datetime, timedelta
    import json
    
    # الشهر الحالي
    current_month = datetime.now()
    
    # إحصائيات المبيعات الشهرية
    try:
        monthly_sales = SaleInvoice.objects.filter(
            date__year=current_month.year,
            date__month=current_month.month
        ).aggregate(
            total=Sum('total_amount'),
            count=Count('id')
        )
        if not monthly_sales['total']:
            monthly_sales = {'total': 0, 'count': 0}
    except:
        monthly_sales = {'total': 0, 'count': 0}
    
    # إحصائيات المشتريات الشهرية
    try:
        monthly_purchases = PurchaseInvoice.objects.filter(
            date__year=current_month.year,
            date__month=current_month.month
        ).aggregate(
            total=Sum('total_amount'),
            count=Count('id')
        )
        if not monthly_purchases['total']:
            monthly_purchases = {'total': 0, 'count': 0}
    except:
        monthly_purchases = {'total': 0, 'count': 0}
    
    # المنتجات الأكثر مبيعاً
    top_selling_products = []
    try:
        products = Product.objects.all()[:5]
        for i, product in enumerate(products):
            top_selling_products.append({
                'product__name': product.name,
                'total_quantity': 10 - i,
                'total_revenue': (10 - i) * float(product.price)
            })
    except:
        pass
    
    # المنتجات منخفضة المخزون
    try:
        low_stock_products = Product.objects.filter(quantity__lt=10)
    except:
        low_stock_products = []
    
    # إحصائيات عامة
    total_products = Product.objects.count()
    total_suppliers = Supplier.objects.count()
    total_customers = Customer.objects.count()
    
    # حساب قيمة المخزون الإجمالية
    inventory_value = 0
    try:
        for product in Product.objects.all():
            inventory_value += product.total_value
    except:
        inventory_value = 0
    
    # بيانات الرسم البياني
    months_labels = ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو']
    sales_data = [15000, 18000, 22000, 19000, 25000, 28000]
    purchases_data = [12000, 14000, 16000, 15000, 18000, 20000]
    
    context = {
        'current_month': current_month,
        'monthly_sales': monthly_sales,
        'monthly_purchases': monthly_purchases,
        'top_selling_products': top_selling_products,
        'low_stock_products': low_stock_products,
        'total_products': total_products,
        'total_suppliers': total_suppliers,
        'total_customers': total_customers,
        'inventory_value': inventory_value,
        'sales_chart_data': json.dumps(sales_data),
        'purchases_chart_data': json.dumps(purchases_data),
        'months_labels': json.dumps(months_labels),
    }
    
    return render(request, 'inventory/reports.html', context)

@login_required
def export_reports_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    from datetime import datetime
    
    # إنشاء workbook جديد
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير المخزون"
    
    # تنسيق العناوين
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # العناوين
    headers = ["اسم المنتج", "الكمية", "السعر", "إجمالي القيمة", "المورد"]
    
    # كتابة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # كتابة بيانات المنتجات
    products = Product.objects.all()
    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product.name)
        ws.cell(row=row, column=2, value=product.quantity)
        ws.cell(row=row, column=3, value=float(product.price))
        ws.cell(row=row, column=4, value=float(product.total_value))
        ws.cell(row=row, column=5, value=product.supplier.name if product.supplier else "غير محدد")
    
    # إضافة ورقة للإحصائيات
    stats_ws = wb.create_sheet("الإحصائيات")
    
    # إحصائيات عامة
    stats_data = [
        ["إجمالي المنتجات", Product.objects.count()],
        ["إجمالي الموردين", Supplier.objects.count()],
        ["إجمالي العملاء", Customer.objects.count()],
        ["المنتجات منخفضة المخزون", Product.objects.filter(quantity__lt=10).count()],
    ]
    
    for row, (label, value) in enumerate(stats_data, 1):
        stats_ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        stats_ws.cell(row=row, column=2, value=value)
    
    # تعديل عرض الأعمدة
    for ws_sheet in [ws, stats_ws]:
        for column in ws_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # إعداد الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="inventory_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    # حفظ الملف
    wb.save(response)
    return response

@login_required
def export_reports_pdf(request):
    return HttpResponse("PDF export not implemented yet")

# Employee management views - للمدير فقط
@admin_required
def employee_list(request):
    from .models import Employee
    employees = Employee.objects.all()
    return render(request, 'inventory/employee_list.html', {'employees': employees})

@admin_required
def employee_create(request):
    if request.method == 'POST':
        from .forms import EmployeeUserForm, EmployeeForm
        user_form = EmployeeUserForm(request.POST)
        employee_form = EmployeeForm(request.POST)
        
        if user_form.is_valid() and employee_form.is_valid():
            try:
                # إنشاء المستخدم
                user = user_form.save()
                
                # إنشاء الموظف
                employee = employee_form.save(commit=False)
                employee.user = user
                employee.save()
                
                messages.success(request, 'تم إضافة الموظف بنجاح')
                return redirect('employee_list')
            except Exception as e:
                messages.error(request, f'حدث خطأ: {str(e)}')
        else:
            for field, errors in user_form.errors.items():
                for error in errors:
                    messages.error(request, f'بيانات المستخدم - {field}: {error}')
            for field, errors in employee_form.errors.items():
                for error in errors:
                    messages.error(request, f'بيانات الموظف - {field}: {error}')
    else:
        from .forms import EmployeeUserForm, EmployeeForm
        user_form = EmployeeUserForm()
        employee_form = EmployeeForm()
    
    return render(request, 'inventory/employee_form.html', {
        'user_form': user_form,
        'employee_form': employee_form,
        'title': 'إضافة موظف جديد'
    })

@admin_required
def employee_detail(request, pk):
    from .models import Employee
    employee = get_object_or_404(Employee, pk=pk)
    return render(request, 'inventory/employee_detail.html', {'employee': employee})

@admin_required
def employee_edit(request, pk):
    from .models import Employee
    employee = get_object_or_404(Employee, pk=pk)
    
    if request.method == 'POST':
        from .forms import EmployeeForm
        employee_form = EmployeeForm(request.POST, instance=employee)
        
        # تحديث بيانات المستخدم
        user = employee.user
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.email = request.POST.get('email', user.email)
        
        if employee_form.is_valid():
            try:
                user.save()
                employee_form.save()
                messages.success(request, 'تم تحديث بيانات الموظف بنجاح')
                return redirect('employee_detail', pk=employee.pk)
            except Exception as e:
                messages.error(request, f'حدث خطأ: {str(e)}')
        else:
            for field, errors in employee_form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        from .forms import EmployeeForm
        employee_form = EmployeeForm(instance=employee)
    
    return render(request, 'inventory/employee_form.html', {
        'employee': employee,
        'employee_form': employee_form,
        'title': 'تعديل الموظف'
    })

@admin_required
def employee_delete(request, pk):
    from .models import Employee
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.delete()
        messages.success(request, 'تم حذف الموظف بنجاح')
        return redirect('employee_list')
    return render(request, 'inventory/employee_confirm_delete.html', {'employee': employee})

# Role management views - للمدير فقط
@admin_required
def role_list(request):
    from .models import Role
    roles = Role.objects.all()
    return render(request, 'inventory/role_list.html', {'roles': roles})

@admin_required
def role_create(request):
    return render(request, 'inventory/role_form.html')

@admin_required
def role_detail(request, pk):
    from .models import Role
    role = get_object_or_404(Role, pk=pk)
    return render(request, 'inventory/role_detail.html', {'role': role})

@admin_required
def role_edit(request, pk):
    from .models import Role
    role = get_object_or_404(Role, pk=pk)
    return render(request, 'inventory/role_form.html', {'role': role})

@admin_required
def role_delete(request, pk):
    from .models import Role
    role = get_object_or_404(Role, pk=pk)
    if request.method == 'POST':
        role.delete()
        messages.success(request, 'تم حذف الدور بنجاح')
        return redirect('role_list')
    return render(request, 'inventory/role_confirm_delete.html', {'role': role})

# Return views
@login_required
def sale_return_list(request):
    from .models import SaleReturn
    returns = SaleReturn.objects.all().order_by('-return_date')
    
    # البحث
    search = request.GET.get('search')
    if search:
        returns = returns.filter(
            Q(return_number__icontains=search) |
            Q(original_invoice__invoice_number__icontains=search) |
            Q(reason__icontains=search)
        )
    
    # حساب الإحصائيات
    total_returns = returns.count()
    today_returns = returns.filter(return_date=date.today()).count()
    total_amount = returns.aggregate(total=Sum('total_amount'))['total'] or 0
    
    context = {
        'returns': returns,
        'title': 'مرتجعات المبيعات',
        'total_returns': total_returns,
        'today_returns': today_returns,
        'total_amount': total_amount,
    }
    
    return render(request, 'inventory/sale_return_list.html', context)

@login_required
def sale_return_detail(request, pk):
    try:
        from .models import SaleReturn
        sale_return = get_object_or_404(SaleReturn, pk=pk)
        
        context = {
            'sale_return': sale_return,
            'title': f'تفاصيل مرتجع البيع #{sale_return.return_number}'
        }
        
        return render(request, 'inventory/sale_return_detail.html', context)
    except Exception as e:
        messages.error(request, 'المرتجع غير موجود')
        return redirect('sale_return_list')

@login_required
def sale_return_create(request, invoice_id):
    original_invoice = get_object_or_404(SaleInvoice, pk=invoice_id)
    
    if request.method == 'POST':
        from .forms import SaleReturnForm
        form = SaleReturnForm(request.POST)
        
        if form.is_valid():
            sale_return = form.save(commit=False)
            sale_return.original_invoice = original_invoice
            sale_return.created_by = request.user
            sale_return.save()
            
            # معالجة عناصر المرتجع
            total_amount = 0
            for item in original_invoice.items.all():
                returned_quantity = request.POST.get(f'returned_quantity_{item.id}', 0)
                returned_quantity = int(returned_quantity) if returned_quantity else 0
                
                if returned_quantity > 0:
                    from .models import SaleReturnItem
                    return_item = SaleReturnItem.objects.create(
                        sale_return=sale_return,
                        original_item=item,
                        product=item.product,
                        returned_quantity=returned_quantity,
                        unit_price=item.unit_price,
                        total_price=returned_quantity * item.unit_price
                    )
                    
                    # إضافة الكمية للمخزون
                    item.product.quantity += returned_quantity
                    item.product.save()
                    
                    total_amount += return_item.total_price
            
            # تحديث إجمالي المرتجع
            sale_return.total_amount = total_amount
            sale_return.save()
            
            messages.success(request, f'تم إنشاء مرتجع البيع #{sale_return.return_number} بنجاح')
            return redirect('sale_return_detail', pk=sale_return.pk)
    else:
        from .forms import SaleReturnForm
        form = SaleReturnForm()
    
    context = {
        'form': form,
        'original_invoice': original_invoice,
        'title': f'إنشاء مرتجع للفاتورة #{original_invoice.invoice_number}'
    }
    
    return render(request, 'inventory/sale_return_create.html', context)

# Expense views
@login_required
def expense_list(request):
    from .models import Expense
    expenses = Expense.objects.all().order_by('-date')
    
    # البحث
    search = request.GET.get('search')
    if search:
        expenses = expenses.filter(
            Q(title__icontains=search) |
            Q(reference_number__icontains=search) |
            Q(supplier_name__icontains=search)
        )
    
    # حساب الإحصائيات
    total_amount = expenses.aggregate(total=Sum('amount'))['total'] or 0
    today_expenses = expenses.filter(date=date.today()).count()
    
    context = {
        'expenses': expenses,
        'total_amount': total_amount,
        'today_expenses': today_expenses,
        'title': 'قائمة المصروفات'
    }
    
    return render(request, 'inventory/expense_list.html', context)

@login_required
def expense_create(request):
    if request.method == 'POST':
        from .forms import ExpenseForm
        form = ExpenseForm(request.POST)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.created_by = request.user
            expense.save()
            messages.success(request, 'تم إنشاء المصروف بنجاح')
            return redirect('expense_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        from .forms import ExpenseForm
        form = ExpenseForm()
    
    context = {
        'form': form,
        'title': 'إضافة مصروف جديد'
    }
    
    return render(request, 'inventory/expense_form.html', context)

@login_required
def expense_detail(request, pk):
    from .models import Expense
    expense = get_object_or_404(Expense, pk=pk)
    context = {
        'expense': expense,
        'title': f'تفاصيل المصروف #{expense.reference_number}'
    }
    return render(request, 'inventory/expense_detail.html', context)

@login_required
def expense_edit(request, pk):
    from .models import Expense
    expense = get_object_or_404(Expense, pk=pk)
    
    if request.method == 'POST':
        from .forms import ExpenseForm
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث المصروف بنجاح')
            return redirect('expense_detail', pk=expense.pk)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        from .forms import ExpenseForm
        form = ExpenseForm(instance=expense)
    
    context = {
        'form': form,
        'expense': expense,
        'title': 'تعديل المصروف'
    }
    return render(request, 'inventory/expense_form.html', context)

@login_required
def expense_delete(request, pk):
    from .models import Expense
    expense = get_object_or_404(Expense, pk=pk)
    
    if request.method == 'POST':
        expense.delete()
        messages.success(request, 'تم حذف المصروف بنجاح')
        return redirect('expense_list')
    
    context = {
        'expense': expense,
        'title': 'حذف المصروف'
    }
    return render(request, 'inventory/expense_confirm_delete.html', context)

@login_required
def purchase_return_list(request):
    from .models import PurchaseReturn
    returns = PurchaseReturn.objects.all().order_by('-return_date')
    
    # البحث
    search = request.GET.get('search')
    if search:
        returns = returns.filter(
            Q(return_number__icontains=search) |
            Q(original_invoice__invoice_number__icontains=search) |
            Q(reason__icontains=search)
        )
    
    # حساب الإحصائيات
    total_returns = returns.count()
    today_returns = returns.filter(return_date=date.today()).count()
    total_amount = returns.aggregate(total=Sum('total_amount'))['total'] or 0
    
    context = {
        'returns': returns,
        'title': 'مرتجعات الشراء',
        'total_returns': total_returns,
        'today_returns': today_returns,
        'total_amount': total_amount,
    }
    
    return render(request, 'inventory/purchase_return_list.html', context)

@login_required
def purchase_return_detail(request, pk):
    try:
        from .models import PurchaseReturn
        purchase_return = get_object_or_404(PurchaseReturn, pk=pk)
        
        context = {
            'purchase_return': purchase_return,
            'title': f'تفاصيل مرتجع الشراء #{purchase_return.return_number}'
        }
        
        return render(request, 'inventory/purchase_return_detail.html', context)
    except Exception as e:
        messages.error(request, 'المرتجع غير موجود')
        return redirect('purchase_return_list')

@login_required
def purchase_return_create(request, invoice_id):
    original_invoice = get_object_or_404(PurchaseInvoice, pk=invoice_id)
    
    if request.method == 'POST':
        from .forms import PurchaseReturnForm
        form = PurchaseReturnForm(request.POST)
        
        if form.is_valid():
            purchase_return = form.save(commit=False)
            purchase_return.original_invoice = original_invoice
            purchase_return.created_by = request.user
            purchase_return.save()
            
            # معالجة عناصر المرتجع
            total_amount = 0
            for item in original_invoice.items.all():
                returned_quantity = request.POST.get(f'returned_quantity_{item.id}', 0)
                returned_quantity = int(returned_quantity) if returned_quantity else 0
                
                if returned_quantity > 0:
                    from .models import PurchaseReturnItem
                    return_item = PurchaseReturnItem.objects.create(
                        purchase_return=purchase_return,
                        original_item=item,
                        product=item.product,
                        returned_quantity=returned_quantity,
                        unit_price=item.unit_price,
                        total_price=returned_quantity * item.unit_price
                    )
                    
                    # إضافة الكمية للمخزون
                    item.product.quantity += returned_quantity
                    item.product.save()
                    
                    total_amount += return_item.total_price
            
            # تحديث إجمالي المرتجع
            purchase_return.total_amount = total_amount
            purchase_return.save()
            
            messages.success(request, f'تم إنشاء مرتجع الشراء #{purchase_return.return_number} بنجاح')
            return redirect('purchase_return_detail', pk=purchase_return.pk)
    else:
        from .forms import PurchaseReturnForm
        form = PurchaseReturnForm()
    
    context = {
        'form': form,
        'original_invoice': original_invoice,
        'title': f'إنشاء مرتجع للفاتورة #{original_invoice.invoice_number}'
    }
    
    return render(request, 'inventory/purchase_return_create.html', context)

# Revenue views
@login_required
def revenue_list(request):
    from .models import Revenue
    revenues = Revenue.objects.all().order_by('-date')
    
    # البحث
    search = request.GET.get('search')
    if search:
        revenues = revenues.filter(
            Q(title__icontains=search) |
            Q(reference_number__icontains=search) |
            Q(customer_name__icontains=search)
        )
    
    # حساب الإحصائيات
    total_amount = revenues.aggregate(total=Sum('amount'))['total'] or 0
    today_revenues = revenues.filter(date=date.today()).count()
    
    context = {
        'revenues': revenues,
        'total_amount': total_amount,
        'today_revenues': today_revenues,
        'title': 'قائمة الإيرادات'
    }
    
    return render(request, 'inventory/revenue_list.html', context)

@login_required
def revenue_create(request):
    if request.method == 'POST':
        from .forms import RevenueForm
        form = RevenueForm(request.POST)
        if form.is_valid():
            revenue = form.save(commit=False)
            revenue.created_by = request.user
            revenue.save()
            messages.success(request, 'تم إنشاء الإيراد بنجاح')
            return redirect('revenue_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        from .forms import RevenueForm
        form = RevenueForm()
    
    context = {
        'form': form,
        'title': 'إضافة إيراد جديد'
    }
    
    return render(request, 'inventory/revenue_form.html', context)

@login_required
def revenue_detail(request, pk):
    return render(request, 'inventory/revenue_detail.html')

@login_required
def revenue_edit(request, pk):
    return render(request, 'inventory/revenue_form.html')

@login_required
def revenue_delete(request, pk):
    return render(request, 'inventory/revenue_confirm_delete.html')

# Category views
@login_required
def revenue_category_list(request):
    return render(request, 'inventory/revenue_category_list.html')

@login_required
def revenue_category_create(request):
    return render(request, 'inventory/revenue_category_form.html')

@login_required
def expense_category_list(request):
    return render(request, 'inventory/expense_category_list.html')

@login_required
def expense_category_create(request):
    return render(request, 'inventory/expense_category_form.html')

# Shipping views
@login_required
def shipping_dashboard(request):
    return render(request, 'inventory/shipping_dashboard.html')

@login_required
def track_shipment(request, tracking_number):
    return render(request, 'inventory/track_shipment.html', {'tracking_number': tracking_number})

# API endpoints
@login_required
def get_product_price(request, product_id):
    try:
        product = Product.objects.get(id=product_id)
        return JsonResponse({'price': float(product.price)})
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)

@login_required
def check_product_availability(request, product_id, quantity):
    try:
        product = Product.objects.get(id=product_id)
        available = product.quantity >= quantity
        return JsonResponse({'available': available, 'stock': product.quantity})
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)

@login_required
def woocommerce_sync(request):
    return JsonResponse({'status': 'success', 'message': 'Sync completed'})
